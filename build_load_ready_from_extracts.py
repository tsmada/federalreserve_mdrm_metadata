import csv
import cProfile
from openpyxl import Workbook

"""Tri-Analytics 2017"""
"""Federal Reserve Board Integration Process"""


class FedAILoader(object):
    """Combine and return workbooks."""

    def __init__(self):
        """Call processing methods on object instantiation."""
        with open('Federal Reserve Systems.csv', 'wb') as f:
            writer = csv.writer(f, delimiter=',',
                            quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(['Division'] + ['TechSystem'] + ['TechGroup'] +
                ['DataSet'] + ['DataElem'] + ['BizTermGroup'] + ['BizTerm'] +
                ['ParentDataElem'] + ['SubType'] + ['SourceUID'] + ['Description'])
        print 'Normalizing'
        self.normalize()
        print 'Normalized'
        self.establish_hierarchy()
        print 'Established Hierarchy'
        self.establish_ref_data()
        print 'Established reference data arrays'
        print 'Preparing to Join references into integrated book'
        self.join()

    def normalize(self):
        self.item_name_desc = []
        self.effective_values = []
        self.expiration_values = []
        with open('data_dict_item_number.csv', 'rU') as f:
            csvreader = csv.reader(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            i=0
            for row in csvreader:
                if i!=0:
                    if row[1].strip() not in self.effective_values:
                        self.effective_values.append(row[1].strip)
                    if row[2].strip() not in self.expiration_values:
                        self.expiration_values.append(row[2].strip())

                    if row[0].strip() + '|' + row[7] not in self.item_name_desc:
                        self.item_name_desc.append(row[0].strip() + '|' + row[7])
                        #print row[0].strip()[4:8]
                i+=1
        return sorted(self.item_name_desc)

    def establish_ref_data(self):
        self.item_desc = []
        self.desc = []
        self.report_desc = []
        self.report_list = []
        with open('data_dict_report_form.csv', 'rU') as f:
            i=0
            for line in f.readlines():
                if i!=0:
                    line = line.replace('"','').strip().split(',')
                    try:
                        if line[3] not in self.desc:
                            self.desc.append(line[3])
                        if line[2] + '|' + line[3] not in self.item_desc:
                            self.item_desc.append(line[2] + '|' + line[3])
                        self.report_desc.append(line[0] + '|' + line[1])
                        if line[0] + ' - ' + line[1] not in self.report_list:
                            self.report_list.append(line[0] + ' - ' + line[1])
                    except IndexError as e:
                        #print 'error somewhere in ref data', line, e
                        pass
                i+=1
        return self.item_desc, self.report_desc, self.report_list


    def establish_hierarchy(self):
        """Use hierarchy from Fed site to build parents"""
        i=0
        with open('data_dict_mdrm_hierarchy.csv', 'rU') as f:
            path = ''
            g = f.readlines()
            path = ''
            parentindex = None
            ultimateParentIndex = None
            self.paths = []
            self.node_desc = []
            for line in g:
                line = line.replace('\n','').replace('"','').split(',')
                #print line
                if i !=0:
                    if len(line) != 3:
                        if line[0] != 'None': ## IF Ultimate Parent
                            ultimateParentIndex = i
                            path = line[0]
                            if {line[0]:line[3]} not in self.node_desc:
                                self.node_desc.append({line[0]:line[3]})
                        if line[1] != 'None': ## IF Sub Series
                            subseries = line[1]
                            up = g[ultimateParentIndex].replace('\n','').replace('"','').split(',')[0]
                            path = up + '/' + line[1]
                            parentindex = i
                            if {line[1]:line[3]} not in self.node_desc:
                                self.node_desc.append({line[1]:line[3]})
                        if line[2] != 'None': ## If Segmented Mnemonic
                            up = g[ultimateParentIndex].replace('\n','').replace('"','').split(',')[0]
                            subseries = g[parentindex].replace('\n','').replace('"','').split(',')[1]
                            path = up + '/' + subseries + '/' + line[2]
                            if {line[2]:line[3]} not in self.node_desc:
                                self.node_desc.append({line[2]:line[3]})
                        if path != '':
                            #print path
                            self.paths.append(path)
                        path = ''
                i+=1
        return self.paths, self.node_desc



    def join(self):
        """Join workbooks together for DataSet grain processing."""
        self.missing_segments = []
        self.ultimate_failures = []
        wb = Workbook(write_only=True)
        self.application_headers = ['Division','System','Application','SourceUID','ParentApplication','Description']
        self.data_elem_headers = ['Division','TechSystem','TechGroup','DataSet','DataElem','ParentDataElem','SubType','SourceUID','Description', 'BizTermGroup','BizTerm', 'BizTermDivision']
        self.data_set_headers = ['Division','TechSystem','TechGroup','DataSet','SubType','ApplicationDivision', 'ApplicationSystem', 'Application', 'SourceUID', 'Description']
        self.biz_term_group_headers = ['Division','BizTermGroup','ParentBizTermGroup']
        self.biz_term_headers = ['Division','BizTermGroup','BizTerm','ParentBizTerm','Description']
        self.application_headers = ['Division','System','Application','SourceUID','ParentApplication','Description']
        self.tagged_object_headers = ['Tag', 'TaggedValue', 'ObjectClass', 'Object']
        self.biz_term_xref_headers = ['BizTermRelationship','ChildDivision','ChildBizTermGroup','ChildBizTerm','ParentDivision','ParentBizTermGroup','ParentBizTerm','Description']
        self.classes_to_write = ['DataElem', 'DataSet', 'BizTermGroup', 'BizTerm', 'Application', 'BizTermXref', 'TaggedObject']
        self.parent_row_created = []
        desubtype = 'ReportField'
        dssubtype = 'Report'
        btr = 'is rationalized by'
        application = 'RISK AND REGULATORY REPORTING/FEDERAL RESERVE REPORTS 2017'
        parentapplication = 'RISK AND REGULATORY REPORTING'
        system = 'WELLS FARGO BANK NATIONAL ASSOCIATION'
        division = 'TRI-ANALYTICS IIP/DEMO-1/WELLS FARGO AND COMPANY/MATERIAL LEGAL ENTITIES'
        techsystem = 'FEDERAL RESERVE REPORTS TECHSYSTEM'
        techgroup = 'FEDERAL RESERVE REPORTS TECHGROUP 2017'
        application = 'RISK AND REGULATORY REPORTING/FEDERAL RESERVE REPORTS 2017'
        btgroup = 'Data Dictionaries/Level 2:  Enterprise Dictionaries/Federal Reserve Report Ontology 2017/' ## + dataset
        btgroup1 = 'Data Dictionaries/Level 2:  Enterprise Dictionaries/Federal Reserve Reporting Business Terms 2017'
        parentbtr = 'Data Dictionaries/Level 2:  Enterprise Dictionaries/Federal Reserve Report Ontology 2017'


        ws = wb.create_sheet(self.classes_to_write[2])
        ws.append(self.biz_term_group_headers)
        ws.append([division, btgroup, 'Data Dictionaries/Level 2:  Enterprise Dictionaries/Federal Reserve Report Ontology 2017', 'Data Dictionaries/Level 2:  Enterprise Dictionaries'])
        ws.append([division, btgroup1, 'Data Dictionaries/Level 2:  Enterprise Dictionaries'])
        for report in self.report_list:
            report = report.replace('/','\\').replace('"','').strip().replace('~',',')
            btgroup = 'Data Dictionaries/Level 2:  Enterprise Dictionaries/Federal Reserve Report Ontology 2017/'
            btgroup = btgroup + str(report)
            ws.append([division, btgroup, parentbtr])
        ws = wb.create_sheet(self.classes_to_write[4])
        ws4 = wb.create_sheet(self.classes_to_write[6])
        ws4.append(self.tagged_object_headers)
        ws.append(self.application_headers)
        ws.append([division, system, application, None, parentapplication])
        ws = wb.create_sheet(self.classes_to_write[3] + "-1")
        ws.append(self.biz_term_headers)
        for term in sorted(self.item_desc):
            term = term.replace('~',',').split('|')
            item = term[0].strip()
            try:
                fulldesc = [desc.split('|')[1] for desc in self.item_name_desc if desc.split('|')[0][4:8] == item][0]
            except:
                fulldesc = ''
                #print [desc for desc in self.item_name_desc if desc.split('|')[0][4:8] == item]
            termd = term[1]
            ws.append([division, btgroup1, termd, None, str(fulldesc)])
        ws = wb.create_sheet(self.classes_to_write[0])
        ws1 = wb.create_sheet(self.classes_to_write[3])
        ws.append(self.data_elem_headers)
        ws1.append(self.biz_term_headers)
        ws2 = wb.create_sheet(self.classes_to_write[5])
        ws2.append(self.biz_term_xref_headers)
        ws3 = wb.create_sheet(self.classes_to_write[5] + '-1')
        ws3.append(self.biz_term_xref_headers)

        ## self.node_desc ==
        ## {'AAXX': 'Risk-Based Capital Reporting for Institutions Subject to the Advanced Capital Adequacy Framework'}


        with open('data_dict_item_number.csv', 'rU') as f: ## Data Element Grain File -- Report Number, Report Name
            reader = csv.reader(f, delimiter=',', quotechar='"')
            i=0
            for line in reader:
                current_assoc = []
                btgroup = 'Data Dictionaries/Level 2:  Enterprise Dictionaries/Federal Reserve Report Ontology 2017/'
                line = [seg.replace('"','').replace('/','\\').replace('~',',') for seg in line]
                #print line
                for j, word in enumerate(line[7].split(' ')):
                    if len(word) >=6:
                        if word[0] == '(' and word[-1] == ')' and word[2:5].isdigit() == True:
                            if word.replace('(','').replace(')','') not in current_assoc:
                                current_assoc.append(word.replace('(','').replace(')',''))
                            # print word.replace('(','').replace(')',''), i ## Returns 5342
                    elif word == 'no.' or word == 'no' or word == 'item' or word == 'items' or word == 'Items':
                        try:
                            if line[7].split(' ')[j+1].replace('.','').isdigit() == True:
                                if line[7].split(' ')[j+1].replace('.','') not in current_assoc:
                                    current_assoc.append(line[7].split(' ')[j+1].replace('.',''))
                                #print line[7].split(' ')[j+1].replace('.',''), i, 'bot'
                        except:
                            pass

                if i!=0:
                    try:
                        try:
                            p = [path for path in self.paths if path.split('/')[-1] == line[0][0:4]][0]
                        except IndexError as e:
                            if line[0][0:4] not in self.missing_segments:
                                self.missing_segments.append(line[0][0:4])
                            p = line[0][0:4]
                        try:
                            desc = line[8]
                        except IndexError as e:
                            desc = ''
                        try:
                            ## something is up with the indexing of this
                            reportdesc = [x for x in sorted(self.report_list) if x.split(' - ')[0].replace('"','').strip().replace('~',',') == line[5].replace('\\','/').replace('"','').replace('~',',').strip()][0].replace('"','')
                            #FR Y-9C - Consolidated Financial Statements for Bank Holding Companies
                        except IndexError as e:
                            reportdesc = ''
                        try:
                            segmentdesc = [x for x in self.node_desc if line[0][0:4] in x][0][line[0][0:4]]
                            #print segmentdesc[line[0][0:4]]
                        except IndexError as e:
                            print 'something went wrong'
                            segmentdesc = ''
                        pathtoelem = p + '/' + line[0] + ' - ' + desc.strip()
                        elemparent = p.split('/')
                        reportdesc = reportdesc.replace('/','\\').replace('~',',').strip()
                        btgroup = btgroup + reportdesc
                        if len(reportdesc) == 0:
                            btgroup = btgroup[:-1]
                        #print btgroup + '\n'
                        if len(elemparent) == 1:
                            #print elemparent[0], segmentdesc
                            if {line[0]:elemparent[0]} not in self.parent_row_created:
                                self.parent_row_created.append({line[0]:elemparent[0]})
                                ws.append([division, techsystem, techgroup, reportdesc, elemparent[0], None,desubtype, None, segmentdesc, btgroup, elemparent[0], division])
                                ws1.append([division, btgroup, elemparent[0], None, segmentdesc])
                        elif len(elemparent) == 2:
                            if {line[0]:pathtoelem} not in self.parent_row_created:
                                #print elemparent, segmentdesc
                                self.parent_row_created.append({line[0]:p})
                                #segmentdesc = [x for x in self.node_desc if elemparent[0] in x][0][elemparent[0]]
                                #print segmentdesc
                                ws.append([division, techsystem, techgroup, reportdesc, p, elemparent[0], desubtype, None, segmentdesc, btgroup, p, division])
                                ws1.append([division, btgroup, p, elemparent[0], segmentdesc])
                                if {reportdesc.split(' - ')[0]:elemparent[0]} not in self.parent_row_created:
                                    currentseg = elemparent[0]
                                    segmentdesc = [x for x in self.node_desc if currentseg in x][0][currentseg]
                                    ws.append([division, techsystem, techgroup, reportdesc, elemparent[0], None, desubtype, None, segmentdesc, btgroup, elemparent[0], division])
                                    ws1.append([division, btgroup, elemparent[0], None, segmentdesc])
                        elif len(elemparent) == 3:
                            # print pathtoelem # BHCF/BHCK/BHDM/BHDM2081 - LOANS TO FOREIGN GOVERNMENTS AND OFFICIAL INSTITUTIONS
                            # print elemparent # ['BHCF', 'BHCK', 'BHDM']
                            # print reportdesc # FR Y-14A - Capital Assessments and Stress Testing
                            # print line[0] # CBPB2081
                            # print p # BHCF/BHCK/BHDM
                            if "and Loan Holding Comapnies" in reportdesc:
                                    print p
                            if {line[0]:pathtoelem} not in self.parent_row_created:
                                if "and Loan Holding Comapnies" in reportdesc:
                                    print p
                                self.parent_row_created.append({line[0]:p})
                                ws.append([division, techsystem, techgroup, reportdesc, '/'.join(elemparent[0:2]), elemparent[0], desubtype, None, segmentdesc, btgroup, '/'.join(elemparent[0:2]), division])
                                ws1.append([division, btgroup,'/'.join(elemparent[0:2]), elemparent[0], segmentdesc])
                            ws.append([division, techsystem, techgroup, reportdesc, p, '/'.join(elemparent[0:2]), desubtype, None, segmentdesc, btgroup, p, division])
                            ws1.append([division, btgroup, p, '/'.join(elemparent[0:2]), segmentdesc])
                        ws.append([division, techsystem, techgroup, reportdesc, pathtoelem, '/'.join(elemparent), desubtype, None, line[7], btgroup, pathtoelem, division])
                        ws2.append([btr, division, btgroup, pathtoelem, division, btgroup1, pathtoelem.split(' - ')[1]])
                        ws1.append([division, btgroup, pathtoelem, '/'.join(elemparent), line[7]])
                        ws4.append(['Effective Date', line[1].replace('\\','/'), 'DataElem', '[' + techsystem + '].[' + techgroup + '].[' + reportdesc + '].[' + pathtoelem + ']'])
                        ws4.append(['Expiration Date', line[2].replace('\\','/'), 'DataElem', '[' + techsystem + '].[' + techgroup + '].[' + reportdesc + '].[' + pathtoelem + ']'])
                        for assoc in current_assoc:
                            try:
                                a = [x.split('|')[1] for x in self.item_desc if x.split('|')[0] == assoc][0]
                                ws3.append([btr, division, btgroup1, a, division, btgroup, pathtoelem])
                            except:
                                pass
                    except IndexError as e:
                        print 'cant expression, ultimate failure', e, i
                        if i not in self.ultimate_failures:
                            self.ultimate_failures.append(i)
                        pass
                i+=1
            ws = wb.create_sheet(self.classes_to_write[1])
            ws.append(self.data_set_headers)
            for x in self.report_list:
                ws.append([division, techsystem, techgroup, x.replace('/','\\').replace('"','').replace('~',',').strip(), dssubtype, division, system, application])


            # ws = wb.create_sheet(self.classes_to_write[3]) ## BizTerm Data Element Grain
            # ws.append(self.biz_term_headers)
            # i=0
        wb.save('Federal Reserve Integrated.xlsx')




if __name__ == '__main__':
    FedAILoader()
